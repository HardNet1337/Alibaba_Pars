from selenium import webdriver
from time import sleep
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup as BS
from selenium.webdriver.common.action_chains import ActionChains
import requests
from openpyxl import Workbook, load_workbook
import os

#Функция get_data задает настройки для драйвера браузера chrome и запускает его
#В качестве значения она принимает ссылку на раздел сайта
#На странице функция скролит до самого низа, ждет одну секунду и снова скролит вниз, таким образом прогружая весь контент
#Внимание! Для корректной работы этой функции в параметре s = Service в свойстве executable_path должен быть прописан путь до драйвера chrome, а его версия должна соотвествовать версии вашего браузера
#Скачать драйвер можно на официальном сайте. Если с этим возникают проблемы, подробную информацию можно найти в документации Selenium
def get_data(url):
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    s = Service(executable_path=r"chromedriver.exe")
    driver = webdriver.Chrome(service=s, options=options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",{
        'source': '''
            delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array
            delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise
            delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol
        '''
        }
    )

    try:
        times = 100
        driver.get(url=url)
        sleep(2)
        bottom = driver.find_element(By.CLASS_NAME, "ui-footer-seo")
        while times > 0:
            print(times)
            action = ActionChains(driver)
            action.move_to_element(to_element=bottom).perform()
            sleep(1)
            times = times - 1

        with open("Ali_page_item.html", 'w', encoding='utf-8') as file:
            file.write(driver.page_source)
            file.close()

    except Exception as ex:
        print("Wrong driver setup")
    finally:
        driver.close()
        driver.quit()

#Функция get_href собирает все товары с прогруженной страницы и сохраняет их в файл
#В качестве значения она принимает путь до html файла страницы, созданом функцией get_data
def get_href(file_path):
    with open(file_path, encoding='utf-8') as file:
        number = 0
        file_links = open('Links.txt', 'a', encoding='utf-8')
        src = file.read()
        soup = BS(src, 'lxml')
        link_case = soup.find('div', class_='hugo4-pc-grid hugo4-pc-grid-5 hugo4-pc-grid-list')
        links = link_case.find_all('div', class_='hugo4-pc-grid-item')
        for items in links:
            try:
                link = items.find_next('a').get('href')
                file_links.write(link + '\n')
                print(f"link {number} collected")
                number = number + 1
            except:
                print('Wrong link data')
        file_links.close()

#Функция get_link_data обрабатывает ссылки, собранные функцией refresh
#Она собирает все заданные атрибуты для товара и записывает их в exel таблицу
def get_link_data():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    s = Service(executable_path=r"chromedriver.exe")
    driver = webdriver.Chrome(service=s, options=options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        'source': '''
               delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array
               delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise
               delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol
           '''
    }
                           )

    file_links = open('Links.txt', 'r', encoding='utf-8')
    check = True
    sub_url = "1"
    number = 1

    while check is True:
        if sub_url != "":
            try:
                sub_url = file_links.readline()

                try:
                    times = 3
                    driver.get(url=sub_url)
                    sleep(2)
                    bottom = driver.find_element(By.CLASS_NAME, "ui-footer-seo")
                    while times > 0:
                        action = ActionChains(driver)
                        action.move_to_element(to_element=bottom).perform()
                        sleep(1)
                        times = times - 1

                    try:
                        driver.find_element(By.CLASS_NAME, "sku-body").find_element(By.CLASS_NAME, "view-more").click()
                    except:
                        print("No 'more' content")

                    with open("Ali_page_item_data.html", 'w', encoding='utf-8') as file:
                        file.write(driver.page_source)
                        file.close()

                except Exception as ex:
                    print("Wrong subdriver setup")

                print("Step 1")

                file_data = open("Ali_page_item_data.html", 'r', encoding='utf-8')
                soup = BS(file_data, 'lxml')
                #item_var определяет тип товара по количеству тэгов 'sku-option'
                item_var_find = []
                item_var_check = False
                try:
                    item_var_first = soup.find('div', class_='sku-body')
                    item_var_find = item_var_first.find('div', class_='sku-option')
                    item_var_len = len(item_var_find)
                except:
                    item_var_len = 0
                if item_var_len > 1:
                    item_var_check = True
                if item_var_check == True:
                    item_var = 'variation'
                    item_tax_class = 'parent'

                else:
                    item_var = 'simple'
                    item_tax_class = ' '

                item_sku_start = sub_url.rfind('_') + 1
                item_sku_end = sub_url.rfind('.')
                item_sku = sub_url[item_sku_start:item_sku_end]

                if item_sku == 'image':
                    item_sku_start = sub_url.find('p-detail')
                    item_sku_end = sub_url.find('.html')
                    item_sku_middle = sub_url[item_sku_start:item_sku_end]
                    item_sku_start = item_sku_middle.rfind('-') + 1
                    item_sku = item_sku_middle[item_sku_start:item_sku_end]

                item_name = soup.find('div', class_='product-title').text
                #item_price определяет цену товара
                #Так как они прописываются двумя разными способами - сначала забирает значение, заданное первым способом, если оно пустое, то - вторым
                item_price_first = soup.find('div', class_='price-list')
                item_price = item_price_first.find('div', class_='price')
                if item_price == None:
                    item_price = item_price_first.find('div', class_='price-range').find('span', class_='price')

                item_categories = soup.find('ul', class_='detail-next-breadcrumb').text
                item_categorie = item_categories.replace('/', '>')
                item_categorie_last = item_categorie.rfind(">")
                item_sort = item_categorie[item_categorie_last + 1:]

                try:
                    item_tag = soup.find('span', class_='hot-sale').text
                except Exception as ex:
                    item_tag = " "

                item_image = " "
                item_image_first = soup.find('div', class_='thumb-list').find('div', class_='detail-next-slick-list')
                if item_var == 'simple':
                    item_image = item_image_first.find('img').get('src')
                else:
                    item_images = item_image_first.find_all('img')
                    for item_image_sort in item_images:
                        item_image_url = item_image_sort.get('src')
                        if item_image:
                            item_image += ","
                        item_image += item_image_url

                item_attributes = {}
                item_attribute_num = 1
                try:
                    item_attribute_first = soup.find('div', class_='sku-body').find_all('div', class_='sku-item')
                    for item_attribute_second in item_attribute_first:
                            try:
                                item_attribute_name = item_attribute_second.find_next('label')
                                item_attribute_value_first = item_attribute_second.find_all('a')
                                item_attribute_value = " "
                                for item_attribute_value_second in item_attribute_value_first:
                                    item_attribute_value = item_attribute_value + item_attribute_value_second.text
                                if item_attribute_value == " ":
                                    item_attribute_value_third = soup.find('div', class_='sku-body').find_all('span', class_='txt')
                                    for item_attribute_values in item_attribute_value_third:
                                        item_attribute_value = item_attribute_value + item_attribute_values.text
                                if item_var == "simple":
                                    item_attribute_visible = "1"
                                    item_attribute_global = "1"
                                if item_var == "variation":
                                    item_attribute_visible = "1"
                                    item_attribute_global = "1"
                                item_attribute = {
                                f"Attribute {item_attribute_num} name": item_attribute_name.text,
                                f"Attribute {item_attribute_num} value(s)": item_attribute_value,
                                f"Attribute {item_attribute_num} visible": item_attribute_visible,
                                f"Attribute {item_attribute_num} global": item_attribute_global
                                }
                                item_attributes.update(item_attribute)
                            except:
                                item_attribute = {
                                    f"Attribute {item_attribute_num} name": " ",
                                    f"Attribute {item_attribute_num} value(s)": " ",
                                    f"Attribute {item_attribute_num} visible": " ",
                                    f"Attribute {item_attribute_num} global": " "
                                    }
                                item_attributes.update(item_attribute)
                            item_attribute_num += 1
                    if item_attribute_num < 10:
                        while item_attribute_num <= 10:
                            item_attribute = {
                                f"Attribute {item_attribute_num} name": " ",
                                f"Attribute {item_attribute_num} value(s)": " ",
                                f"Attribute {item_attribute_num} visible": " ",
                                f"Attribute {item_attribute_num} global": " "
                            }
                            item_attributes.update(item_attribute)
                            item_attribute_num += 1

                except:
                    if item_attribute_num < 10:
                        while item_attribute_num <= 10:
                            item_attribute = {
                                f"Attribute {item_attribute_num} name": " ",
                                f"Attribute {item_attribute_num} value(s)": " ",
                                f"Attribute {item_attribute_num} visible": " ",
                                f"Attribute {item_attribute_num} global": " "
                            }
                            item_attributes.update(item_attribute)
                            item_attribute_num += 1

                item_description_overviev1 = " "
                item_description_desc1 = " "
                try:
                    item_description_overviev = soup.find('div', class_='do-content')
                    item_description_overviev1 = str(item_description_overviev.prettify())
                except Exception as exDesc:
                    item_description_overviev = " "
                try:
                    item_description_desc = soup.find('div', class_='ife-detail-decorate-table')
                    item_description_desc1 = str(item_description_desc.prettify())
                except Exception as exDesc:
                    item_description_desc = " "
                item_description = f"{item_description_desc}"
                item_description_short = item_description_overviev
                if item_description == " ":
                    try:
                        item_description_overviev = soup.find('div', class_='do-entry do-entry-separate')
                        item_description_overviev1 = str(item_description_overviev.prettify())
                    except:
                        item_description_overviev = " "
                if item_description_short == " ":
                    try:
                        item_description_desc = soup.find('div', class_='aliDataTable')
                        item_description_desc1 = str(item_description_desc.prettify())
                    except :
                        item_description_desc = " "
                    item_description = f"{item_description_overviev, item_description_desc}"
                    item_description_short = item_description_overviev

                try:
                    item_description_main_short = soup.find('div', class_='do-entry-list')
                    item_description_short = " "
                    item_descriptions_short = item_description_main_short.find_all('dl')
                    for item_desc_item in item_descriptions_short:
                        item_description_short = item_description_short + item_desc_item.text + "<br/>"
                except:
                    item_description_short = " "

                item_stock_first = soup.find('div', class_='lead-list').find('tr').find_all('td')
                item_stock_second = item_stock_first[1].text
                item_stock_third = item_stock_second.split('-')
                item_stock = int(item_stock_third[1].strip())
                if item_stock > 0:
                    item_in_stock = "1"
                else:
                    item_in_stock = "0"

                try:
                    item_seller_url = soup.find('a', class_='company-name company-name-lite-vb').get('href')
                except:
                    item_seller_url = soup.find('div', class_='company-head').find('a').get('href')

                try:
                    item_rate_block = soup.find('div', class_='review-conclusion').find('span', class_='next-form-text-align review-value').text
                except:
                    item_rate_block = " "
                item_rate = item_rate_block

                #Это список атрибутов, значения которых всегда статичны, кроме url товара
                item_published = '1'
                item_feat = '0'
                item_visiblity = 'visible'
                item_tax = 'taxable'
                item_reviews = '0'
                item_parent = 'st2'
                item_upsells = 'st2'
                item_crosssells = 'st2'
                item_url = sub_url
                item_ean = '-'

                print("Step 2")

                item_data = {
                    "Type": item_var,
                    "SKU": item_sku,
                    "Name": item_name,
                    "Published": item_published,
                    "Is featured": item_feat,
                    "Visibility in catalogue": item_visiblity,
                    "Short description": item_description,
                    "Description": item_description_short,
                    "Tax status": item_tax,
                    "Tax class": item_tax_class,
                    "In Stock?": item_in_stock,
                    "Stock": item_stock,
                    "Weight(g)": " ",
                    "Length(cm)": " ",
                    "Width(cm)":" ",
                    "Height(cm)":" ",
                    "Allow customer revievs?":item_reviews,
                    "Sale price":item_price.text,
                    "Regular price":item_price.text,
                    "Cetegories":item_categorie,
                    "Tags":item_tag,
                    "Images":item_image,
                    "Parent":item_parent,
                    "Upsells":item_upsells,
                    "Cross-sells":item_crosssells,
                    "External URL":item_url,
                    "Meta: _pris": item_price.text,
                    "Meta: _ean_code": item_ean,
                    "Attribute 1 name": item_attributes[f"Attribute 1 name"],
                    "Attribute 1 value(s)": item_attributes[f"Attribute 1 value(s)"],
                    "Attribute 1 visible": item_attributes[f"Attribute 1 visible"],
                    "Attribute 1 global": item_attributes[f"Attribute 1 global"],
                    "Attribute 2 name": item_attributes[f"Attribute 2 name"],
                    "Attribute 2 value(s)": item_attributes[f"Attribute 2 value(s)"],
                    "Attribute 2 visible": item_attributes[f"Attribute 2 visible"],
                    "Attribute 2 global": item_attributes[f"Attribute 2 global"],
                    "Attribute 3 name": item_attributes[f"Attribute 3 name"],
                    "Attribute 3 value(s)": item_attributes[f"Attribute 3 value(s)"],
                    "Attribute 3 visible": item_attributes[f"Attribute 3 visible"],
                    "Attribute 3 global": item_attributes[f"Attribute 3 global"],
                    "Attribute 4 name": item_attributes[f"Attribute 4 name"],
                    "Attribute 4 value(s)": item_attributes[f"Attribute 4 value(s)"],
                    "Attribute 4 visible": item_attributes[f"Attribute 4 visible"],
                    "Attribute 4 global": item_attributes[f"Attribute 4 global"],
                    "Attribute 5 name": item_attributes[f"Attribute 5 name"],
                    "Attribute 5 value(s)": item_attributes[f"Attribute 5 value(s)"],
                    "Attribute 5 visible": item_attributes[f"Attribute 5 visible"],
                    "Attribute 5 global": item_attributes[f"Attribute 5 global"],
                    "Attribute 6 name": item_attributes[f"Attribute 6 name"],
                    "Attribute 6 value(s)": item_attributes[f"Attribute 6 value(s)"],
                    "Attribute 6 visible": item_attributes[f"Attribute 6 visible"],
                    "Attribute 6 global": item_attributes[f"Attribute 6 global"],
                    "Attribute 7 name": item_attributes[f"Attribute 7 name"],
                    "Attribute 7 value(s)": item_attributes[f"Attribute 7 value(s)"],
                    "Attribute 7 visible": item_attributes[f"Attribute 7 visible"],
                    "Attribute 7 global": item_attributes[f"Attribute 7 global"],
                    "Attribute 8 name": item_attributes[f"Attribute 8 name"],
                    "Attribute 8 value(s)": item_attributes[f"Attribute 8 value(s)"],
                    "Attribute 8 visible": item_attributes[f"Attribute 8 visible"],
                    "Attribute 8 global": item_attributes[f"Attribute 8 global"],
                    "Attribute 9 name": item_attributes[f"Attribute 9 name"],
                    "Attribute 9 value(s)": item_attributes[f"Attribute 9 value(s)"],
                    "Attribute 9 visible": item_attributes[f"Attribute 9 visible"],
                    "Attribute 9 global": item_attributes[f"Attribute 9 global"],
                    "Attribute 10 name": item_attributes[f"Attribute 10 name"],
                    "Attribute 10 value(s)": item_attributes[f"Attribute 10 value(s)"],
                    "Attribute 10 visible": item_attributes[f"Attribute 10 visible"],
                    "Attribute 10 global": item_attributes[f"Attribute 10 global"],
                    "Meta: _m_link": item_seller_url,
                    "Item rate": item_rate
                }
                print(f"Item {number} collected")
                number = number + 1
                exel_write('Items', item_data=item_data)
                print("Step 3")
            except Exception as ex:
                log = open('Items_log.txt', 'w', encoding='utf-8')
                log_data = str(ex) + '\n' + '_' * 60 + '\n'
                log.writelines(log_data)
                log.close()
        else:
            check = False
    file_links.close()
    file_data.close()
    driver.close()
    driver.quit()
    print("Collecting data is finished")

#Функция refresh задает исходный файл для функции get_data, чистит файл для сохранения ссылок
def refresh(file_name):
    file = open(f'{file_name}.txt', 'r', encoding='utf-8')
    file_clean = open('Links.txt', 'w', encoding='utf-8')
    file_clean.write('')
    file_clean.close()
    check = True
    url = "1"
    while check is True:
        if url != "":
            try:
                url = file.readline()
                get_data(url=url)
                get_href('Ali_page_item.html')
            except Exception as ex:
                print(ex)
        else:
            check = False
            file.close()

#exel_write принимает значение всех атрибутов товара и записывает их в соотвествующий exel-файл
def exel_write(file_name, item_data):
    file = f"{file_name}.xlsx"
    if os.path.exists(file):
        wb = load_workbook(file)
        ws = wb.active
        ws.append([item_data["Type"],
        item_data["SKU"],
        item_data["Name"],
        item_data["Published"],
        item_data["Is featured"],
        item_data["Visibility in catalogue"],
        item_data["Short description"],
        item_data["Description"],
        item_data["Tax status"],
        item_data["Tax class"],
        item_data["In Stock?"],
        item_data["Stock"],
        item_data["Weight(g)"],
        item_data["Length(cm)"],
        item_data["Width(cm)"],
        item_data["Height(cm)"],
        item_data["Allow customer revievs?"],
        item_data["Sale price"],
        item_data["Regular price"],
        item_data["Cetegories"],
        item_data["Tags"],
        item_data["Images"],
        item_data["Parent"],
        item_data["Upsells"],
        item_data["Cross-sells"],
        item_data["External URL"],
        item_data["Meta: _pris"],
        item_data["Meta: _ean_code"],
        item_data["Attribute 1 name"],
        item_data["Attribute 1 value(s)"],
        item_data["Attribute 1 visible"],
        item_data["Attribute 1 global"],
        item_data["Attribute 2 name"],
        item_data["Attribute 2 value(s)"],
        item_data["Attribute 2 visible"],
        item_data["Attribute 2 global"],
        item_data["Attribute 3 name"],
        item_data["Attribute 3 value(s)"],
        item_data["Attribute 3 visible"],
        item_data["Attribute 3 global"],
        item_data["Attribute 4 name"],
        item_data["Attribute 4 value(s)"],
        item_data["Attribute 4 visible"],
        item_data["Attribute 4 global"],
        item_data["Attribute 5 name"],
        item_data["Attribute 5 value(s)"],
        item_data["Attribute 5 visible"],
        item_data["Attribute 5 global"],
        item_data["Attribute 6 name"],
        item_data["Attribute 6 value(s)"],
        item_data["Attribute 6 visible"],
        item_data["Attribute 6 global"],
        item_data["Attribute 7 name"],
        item_data["Attribute 7 value(s)"],
        item_data["Attribute 7 visible"],
        item_data["Attribute 7 global"],
        item_data["Attribute 8 name"],
        item_data["Attribute 8 value(s)"],
        item_data["Attribute 8 visible"],
        item_data["Attribute 8 global"],
        item_data["Attribute 9 name"],
        item_data["Attribute 9 value(s)"],
        item_data["Attribute 9 visible"],
        item_data["Attribute 9 global"],
        item_data["Attribute 10 name"],
        item_data["Attribute 10 value(s)"],
        item_data["Attribute 10 visible"],
        item_data["Attribute 10 global"],
        item_data["Meta: _m_link"],
        item_data["Item rate"]
                   ])
    else:
        with open("Items_test", "w", encoding="utf-8"):
            wb = Workbook()
            ws = wb.active
            ws.append(["Type", "SKU", "Name", "Published", "Is featured", "Visibility in catalogue", "Short description",
               "Description", "Tax status", "Tax class", "In Stock?", "Stock", "Weight(g)", "Length(cm)", "Width(cm)",
               "Height(cm)", "Allow customer revievs?", "Sale price", "Regular price", "Cetegories", "Tags", "Images",
               "Parent", "Upsells", "Cross-sells", "External URL", "Meta: _pris", "Meta: _ean_code", "Attribute 1 name",
               "Attribute 1 value(s)", "Attribute 1 visible", "Attribute 1 global", "Attribute 2 name",
               "Attribute 2 value(s)", "Attribute 2 visible", "Attribute 2 global", "Attribute 3 name",
               "Attribute 3 value(s)", "Attribute 3 visible", "Attribute 3 global", "Attribute 4 name",
               "Attribute 4 value(s)", "Attribute 4 visible", "Attribute 4 global", "Attribute 5 name",
               "Attribute 5 value(s)", "Attribute 5 visible", "Attribute 5 global", "Attribute 6 name",
               "Attribut 6 value(s)", "Attribute 6 visible", "Attribute 6 global", "Attribute 7 name",
               "Attribute 7 value(s)", "Attribute 7 visible", "Attribute 7 global", "Attribute 8 name",
               "Attribute 8 value(s)", "Attribute 8 visible", "Attribute 8 global", "Attribute 9 name",
               "Attribute 9 value(s)", "Attribute 9 visible", "Attribute 9 global", "Attribute 10 name",
               "Attribute 10 value(s)", "Attribute 10 visible", "Attribute 10 global", "Meta: _m_link", "Item rate"])
    wb.save(file)

def exel_remove():
    print("Start removing exel")
    file_name1 = "Items"
    file_name2 = "Items.xlsx"
    if os.path.exists(file_name1):
        os.remove(file_name1)
    if os.path.exists(file_name2):
        os.remove(file_name2)
    print("Exel removed")

def main():
    exel_remove()
    refresh("Categories")
    get_link_data()

if __name__ == "__main__":
    main()